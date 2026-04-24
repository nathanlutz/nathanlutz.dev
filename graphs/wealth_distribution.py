"""
US Wealth Distribution, 1945–2019
Mean net worth per adult by wealth group, animated by year.

Data: Piketty, Saez & Zucman (2022) — Appendix Tables II (Distributional)
      https://gabriel-zucman.eu/usdina/
Units: Real 2019 USD, equal-split adults 20+
Output: public/graphs/wealth_distribution.gif
"""

import os
import io
import warnings
import urllib.request

import numpy as np
import matplotlib.pyplot as plt
import matplotlib.animation as animation
import openpyxl

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

PSZ_CACHE = os.path.join(os.path.dirname(__file__), "psz2022.xlsx")

if not os.path.exists(PSZ_CACHE):
    print("Downloading PSZ 2022 data…")
    url = "https://gabriel-zucman.eu/files/PSZ2022AppendixTablesII(Distrib).xlsx"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req) as r:
        data = r.read()
    with open(PSZ_CACHE, "wb") as f:
        f.write(data)

wb = openpyxl.load_workbook(PSZ_CACHE, read_only=True, data_only=True)

# avghweal sheet — col 1: mean wealth of all adults (avghweal0indiv)
avg_by_year = {
    int(row[0]): float(row[1])
    for row in wb["avghweal"].iter_rows(values_only=True)
    if isinstance(row[0], (int, float)) and row[1] is not None
}

# TE1 sheet — wealth shares by group, rows 9+ are data
# Col indices: [1] Bot90  [4] Top10  [6] Top1  [8] Top0.1  [9] Top0.01
te1_by_year = {}
for row in list(wb["TE1"].iter_rows(values_only=True))[9:]:
    yr = row[0]
    if not isinstance(yr, (int, float)) or row[1] is None:
        continue
    te1_by_year[int(yr)] = {
        "bot90":  row[1],
        "top10":  row[4],
        "top1":   row[6],
        "top01":  row[8],
        "top001": row[9],
    }

# ---------------------------------------------------------------------------
# Wealth groups
# Formula: mean_group = avg_all_adults × share_group / fraction_of_population
# ---------------------------------------------------------------------------

GROUPS = [
    # x_pos  share_fn                               pop_fraction
    (1,  lambda s: s["bot90"],                      0.9000),
    (2,  lambda s: s["top10"] - s["top1"],          0.0900),
    (3,  lambda s: s["top1"]  - s["top01"],         0.0090),
    (4,  lambda s: s["top01"],                      0.0010),
    (5,  lambda s: s["top001"],                     0.0001),
]

def compute_year(year):
    avg0   = avg_by_year.get(year)
    shares = te1_by_year.get(year)
    if avg0 is None or shares is None:
        return None
    pts = []
    for x_pos, share_fn, frac in GROUPS:
        share = share_fn(shares)
        if share is not None and share > 0:
            pts.append((x_pos, avg0 * share / frac))
    return pts or None

plot_data = {yr: pts for yr in range(1945, 2020) if (pts := compute_year(yr))}
years     = sorted(plot_data)

# ---------------------------------------------------------------------------
# Animation helpers
# ---------------------------------------------------------------------------

X_TICKS  = [1,       2,       3,        4,         5]
X_LABELS = ["90th", "99th", "99.9th", "99.99th", "99.999th"]

def smooth_curve(pts):
    xs = np.array([0] + [p[0] for p in pts])
    ys = np.array([0.0] + [p[1] for p in pts])
    x_smooth = np.linspace(0, xs[-1], 500)
    return x_smooth, np.interp(x_smooth, xs, ys)

def smooth_ylims(ymaxes, window=3):
    n = len(ymaxes)
    return [max(ymaxes[max(0, i-window):min(n, i+window+1)]) for i in range(n)]

raw_ymaxes = [max(y for _, y in plot_data[yr]) * 1.08 for yr in years]
ylims      = smooth_ylims(raw_ymaxes)

# ---------------------------------------------------------------------------
# Build and save GIF
# ---------------------------------------------------------------------------

fig, ax = plt.subplots(figsize=(13, 7))
plt.subplots_adjust(left=0.13, right=0.97, top=0.88, bottom=0.13)

line,     = ax.plot([], [], color="#e63946", linewidth=2.5, zorder=3)
dots,     = ax.plot([], [], "o", color="#e63946", markersize=5, zorder=4)
yr_label  = ax.text(0.97, 0.95, "", transform=ax.transAxes,
                    fontsize=40, fontweight="bold", ha="right", va="top",
                    color="#231f20", alpha=0.2)

ax.set_xlim(0.5, 5.5)
ax.set_ylim(0, ylims[0])
ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"${x/1e6:.1f}M"))
ax.set_xticks(X_TICKS)
ax.set_xticklabels(X_LABELS, fontsize=10)
ax.set_xlabel("Percentile (mean wealth of each group)", fontsize=11)
ax.set_ylabel("Mean net worth per adult (real 2019 USD)", fontsize=11)
ax.set_title(
    "Mean Net Worth by Wealth Group, USA 1945–2019\n"
    "Source: Piketty, Saez & Zucman (2022) — real 2019 dollars, equal-split adults 20+",
    fontsize=12,
)
ax.grid(True, alpha=0.3)

def update(frame):
    pts = plot_data[years[frame]]
    xs_s, ys_s = smooth_curve(pts)
    line.set_data(xs_s, ys_s)
    dots.set_data([p[0] for p in pts], [p[1] for p in pts])
    yr_label.set_text(str(years[frame]))
    ax.set_ylim(0, ylims[frame])
    return line, dots, yr_label

ani = animation.FuncAnimation(
    fig, update, frames=len(years),
    interval=700, blit=False, repeat=True, repeat_delay=1500,
)

out = os.path.join(os.path.dirname(__file__), "..", "public", "graphs", "wealth_distribution.gif")
ani.save(out, writer="pillow", fps=1.3, dpi=120)
print(f"Saved {out}  ({len(years)} frames, {years[0]}–{years[-1]})")
